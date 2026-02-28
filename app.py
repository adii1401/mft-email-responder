import os
import requests
import streamlit as st
from groq import Groq
from sample_emails import past_emails
from dotenv import load_dotenv
import msal
import re
import openpyxl
from docx import Document
import pdfplumber
import chromadb
from chromadb.utils import embedding_functions

load_dotenv()

st.set_page_config(page_title="MFT Email Responder", page_icon="📧", layout="wide")

client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# ─── Azure / Graph API Config ────────────────────────────────────
CLIENT_ID = os.environ.get("AZURE_CLIENT_ID")
TENANT_ID = os.environ.get("AZURE_TENANT_ID")
SCOPES = ["Mail.ReadWrite", "Mail.Send", "User.Read"]
GRAPH_ENDPOINT = "https://graph.microsoft.com/v1.0"

# ─── ChromaDB Setup ──────────────────────────────────────────────
@st.cache_resource
def setup_chromadb():
    chroma_client = chromadb.PersistentClient(path="./chroma_db")
    ef = embedding_functions.DefaultEmbeddingFunction()
    collection = chroma_client.get_or_create_collection(
        name="mft_docs",
        embedding_function=ef
    )
    return chroma_client, collection, ef

chroma_client, collection, embedding_model = setup_chromadb()

# ─── Doc readers ────────────────────────────────────────────────
def read_txt(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def read_docx(path):
    doc = Document(path)
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

def read_xlsx(path):
    wb = openpyxl.load_workbook(path)
    result = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join([str(c) for c in row if c is not None])
            if row_text.strip():
                result.append(row_text)
    return "\n".join(result)

def read_pdf(path):
    text = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text.append(t)
    return "\n".join(text)

# ─── Chunking ────────────────────────────────────────────────────
def chunk_text(text, chunk_size=500, overlap=50):
    words = text.split()
    chunks = []
    i = 0
    while i < len(words):
        chunk = " ".join(words[i:i+chunk_size])
        chunks.append(chunk)
        i += chunk_size - overlap
    return chunks

# ─── Load emails into ChromaDB ───────────────────────────────────
def load_emails_into_chromadb(collection):
    existing = collection.get(where={"type": "email"})
    if existing["ids"]:
        return len(existing["ids"])
    documents, ids, metadatas = [], [], []
    for i, email in enumerate(past_emails):
        text = f"Query: {email['query']}\nReply: {email['reply']}"
        documents.append(text)
        ids.append(f"email_{i}")
        metadatas.append({"type": "email", "source": "past_emails"})
    collection.add(documents=documents, ids=ids, metadatas=metadatas)
    return len(documents)

# ─── Load docs into ChromaDB ─────────────────────────────────────
def load_docs_into_chromadb(collection, folder="docs"):
    if not os.path.exists(folder):
        return 0
    existing = collection.get(where={"type": "doc"})
    if existing["ids"]:
        return len(existing["ids"])
    all_chunks, all_ids, all_metadata = [], [], []
    chunk_id = 0
    for filename in os.listdir(folder):
        path = os.path.join(folder, filename)
        try:
            if filename.endswith(".txt"):
                text = read_txt(path)
            elif filename.endswith(".docx"):
                text = read_docx(path)
            elif filename.endswith(".xlsx"):
                text = read_xlsx(path)
            elif filename.endswith(".pdf"):
                text = read_pdf(path)
            else:
                continue
            chunks = chunk_text(text)
            for chunk in chunks:
                if chunk.strip():
                    all_chunks.append(chunk)
                    all_ids.append(f"chunk_{chunk_id}")
                    all_metadata.append({"type": "doc", "source": filename})
                    chunk_id += 1
        except Exception as e:
            st.warning(f"Error reading {filename}: {e}")
    if all_chunks:
        collection.add(documents=all_chunks, ids=all_ids, metadatas=all_metadata)
    return len(all_chunks)

# ─── Confidence scoring ──────────────────────────────────────────
def distance_to_confidence(distance):
    confidence = max(0.0, 1.0 - (distance / 2.0)) * 100
    return round(confidence, 1)

def confidence_badge(score):
    if score >= 70:
        return f"🟢 {score}% match"
    elif score >= 40:
        return f"🟡 {score}% match"
    else:
        return f"🔴 {score}% match"

# ─── Semantic search ─────────────────────────────────────────────
def search_docs(collection, query, top_k=5):
    results = collection.query(
        query_texts=[query],
        n_results=top_k,
        include=["documents", "metadatas", "distances"]
    )
    chunks = results["documents"][0]
    metadatas = results["metadatas"][0]
    distances = results["distances"][0]
    sources = [m["source"] for m in metadatas]
    types = [m["type"] for m in metadatas]
    confidences = [distance_to_confidence(d) for d in distances]
    return chunks, sources, types, confidences

# ─── Generate reply ──────────────────────────────────────────────
def generate_reply(email_text, collection):
    relevant_chunks, sources, types, confidences = search_docs(collection, email_text)
    retrieved_context = "\n\n".join([
        f"[{t.upper()} | Source: {src} | Confidence: {conf}%]\n{chunk}"
        for chunk, src, t, conf in zip(relevant_chunks, sources, types, confidences)
    ])
    top_confidence = confidences[0] if confidences else 0
    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": f"""You are an MFT/EDI support assistant. Use the following retrieved knowledge base context to draft a professional reply.

RETRIEVED CONTEXT (semantically matched to this query):
{retrieved_context}

Now draft a professional reply to this new email:
{email_text}

Important: Follow approval rules strictly. If password reset is requested, mention JO approval is needed and CC the Job Owner.

Format the reply as a proper email with clear paragraph breaks. Do not merge everything into one paragraph.
Do NOT include a Subject line in your reply. Start directly with the greeting."""}]
    )
    reply = response.choices[0].message.content
    # Strip any Subject line the LLM adds at the top
    reply = "\n".join([line for line in reply.split("\n") if not line.strip().startswith("Subject:")])
    return reply.strip(), relevant_chunks, sources, types, confidences, top_confidence

# ─── Microsoft Graph OAuth ───────────────────────────────────────
def get_access_token():
    if "access_token" in st.session_state:
        return st.session_state["access_token"]
    return None

def start_oauth_flow():
    app = msal.PublicClientApplication(CLIENT_ID, authority="https://login.microsoftonline.com/consumers")
    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        st.error("Failed to create device flow. Check your Azure App Registration.")
        return None, None
    st.session_state["msal_flow"] = flow
    st.session_state["msal_app"] = app
    return flow["user_code"], flow["verification_uri"]

def complete_oauth_flow():
    if "msal_flow" not in st.session_state:
        return None
    app = st.session_state["msal_app"]
    flow = st.session_state["msal_flow"]
    result = app.acquire_token_by_device_flow(flow)
    if "access_token" in result:
        st.session_state["access_token"] = result["access_token"]
        return result["access_token"]
    else:
        st.error(f"Auth failed: {result.get('error_description', 'Unknown error')}")
        return None

# ─── Fetch unread emails ─────────────────────────────────────────
def fetch_unread_emails(access_token, max_emails=5):
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"{GRAPH_ENDPOINT}/me/mailfolders/inbox/messages"
    params = {
        "$filter": "isRead eq false",
        "$select": "subject,body,from,receivedDateTime,id",
        "$top": max_emails,
        "$orderby": "receivedDateTime desc"
    }
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        return response.json().get("value", [])
    else:
        st.error(f"Graph API error {response.status_code}: {response.text}")
        return []

# ─── Send reply via Graph API ────────────────────────────────────
def send_reply(access_token, email_id, reply_text, to_address, subject):
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    # Extract CC email from reply text if present
    cc_recipients = []
    for line in reply_text.split("\n"):
        if line.strip().lower().startswith("cc:"):
            import re
            emails_found = re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', line)
            for em in emails_found:
                cc_recipients.append({"emailAddress": {"address": em}})

    url = f"{GRAPH_ENDPOINT}/me/messages/{email_id}/reply"
    body = {"comment": reply_text}

    # If CC found, use createReply + send instead
    if cc_recipients:
        # Step 1: Create a draft reply
        create_url = f"{GRAPH_ENDPOINT}/me/messages/{email_id}/createReply"
        draft_resp = requests.post(create_url, headers=headers)
        if draft_resp.status_code != 201:
            st.write(f"Draft error: {draft_resp.status_code} {draft_resp.text}")
            return False

        draft_id = draft_resp.json().get("id")

        # Step 2: Update draft with body + CC
        update_url = f"{GRAPH_ENDPOINT}/me/messages/{draft_id}"
        update_body = {
            "body": {
                "contentType": "HTML",
                "content": reply_text.replace("\n", "<br>")
            },
            "ccRecipients": cc_recipients
        }
        requests.patch(update_url, headers=headers, json=update_body)

        # Step 3: Send the draft
        send_url = f"{GRAPH_ENDPOINT}/me/messages/{draft_id}/send"
        send_resp = requests.post(send_url, headers=headers)
        return send_resp.status_code == 202
    else:
        response = requests.post(url, headers=headers, json=body)
        if response.status_code != 202:
            st.write(f"Debug — Status: {response.status_code}, Response: {response.text}")
        return response.status_code == 202
def mark_as_read(access_token, email_id):
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }
    url = f"{GRAPH_ENDPOINT}/me/messages/{email_id}"
    requests.patch(url, headers=headers, json={"isRead": True})
    
# ─── Load data on startup ────────────────────────────────────────
total_emails = load_emails_into_chromadb(collection)
total_doc_chunks = load_docs_into_chromadb(collection, "docs")

# ─── Streamlit UI ────────────────────────────────────────────────
st.title("📧 MFT Support Email Responder")
st.markdown("AI-powered draft replies for MFT/EDI support emails using RAG + LLaMA 3.3")

with st.sidebar:
    st.header("📂 Knowledge Base")
    st.markdown(f"**{len(past_emails)} past emails** indexed in ChromaDB 🧠")
    st.markdown(f"**{total_doc_chunks} doc chunks** indexed in ChromaDB 🧠")
    with st.expander("📄 Docs loaded"):
        if os.path.exists("docs") and os.listdir("docs"):
            for f in os.listdir("docs"):
                st.markdown(f"✅ {f}")
        else:
            st.markdown("No docs found in docs/ folder")
    for i, e in enumerate(past_emails, 1):
        with st.expander(f"Email {i}"):
            st.markdown(f"**Query:** {e['query']}")
            st.markdown(f"**Reply:** {e['reply']}")

tab1, tab2 = st.tabs(["📬 Outlook Inbox (Auto)", "✍️ Manual Input"])

# ══════════════════════════════════════════════════════════════════
# TAB 1 — Outlook Auto Mode
# ══════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("📬 Auto-fetch from Outlook Inbox")
    access_token = get_access_token()

    if not access_token:
        st.info("Connect your Outlook account to auto-fetch unread MFT support emails.")
        if st.button("🔐 Connect Outlook Account"):
            user_code, verification_uri = start_oauth_flow()
            if user_code:
                st.session_state["oauth_started"] = True
                st.success(f"**Step 1:** Go to [{verification_uri}]({verification_uri})")
                st.code(user_code, language=None)
                st.info("**Step 2:** Enter the code above → sign in with mft_test_ag1@outlook.com → come back here")

        if st.session_state.get("oauth_started") and not access_token:
            if st.button("✅ I've completed sign-in — fetch my emails"):
                with st.spinner("Completing authentication..."):
                    token = complete_oauth_flow()
                    if token:
                        st.success("Connected! Fetching emails...")
                        st.rerun()
    else:
        st.success("✅ Connected to Outlook")
        col1, col2 = st.columns([1, 4])
        with col1:
            max_emails = st.selectbox("Fetch last", [3, 5, 10], index=1)
        with col2:
            if st.button("🔄 Refresh Inbox"):
                st.rerun()

        with st.spinner("Fetching unread emails from Outlook..."):
            emails = fetch_unread_emails(access_token, max_emails)

        if not emails:
            st.info("No unread emails found. Send a test MFT email to mft_test_ag1@outlook.com and refresh.")
        else:
            st.markdown(f"**{len(emails)} unread email(s) found:**")
            for email in emails:
                subject = email.get("subject", "No Subject")
                sender = email.get("from", {}).get("emailAddress", {}).get("address", "Unknown")
                preview = email.get("bodyPreview", "")
                email_id = email.get("id")
                received = email.get("receivedDateTime", "")[:10]

                with st.expander(f"📩 {subject} — from {sender} ({received})"):
                    body_content = email.get("body", {}).get("content", preview)
                    body_clean_preview = re.sub(r'<[^>]+>', ' ', body_content).strip()[:300]
                    st.markdown(f"**From:** {sender}")
                    st.markdown(f"**Preview:** {body_clean_preview if body_clean_preview else '_(no preview available)_'}")
                    
                    st.divider()

                    # ─── Check if already sent ───────────────────
                    sent_key = f"sent_{email_id}"
                    reply_key = f"reply_{email_id}"

                    if st.session_state.get(sent_key):
                        st.success("✅ Reply already sent for this email.")
                    else:
                        if st.button(f"⚡ Generate Reply", key=f"gen_{email_id}"):
                            with st.spinner("Running RAG + generating reply..."):
                                body_content = email.get("body", {}).get("content", preview)
                                body_clean = re.sub(r'<[^>]+>', ' ', body_content).strip()
                                email_text = f"From: {sender}\n\n{body_clean}"
                                reply, chunks, sources, types, confidences, top_conf = generate_reply(email_text, collection)
                                st.session_state[reply_key] = {
                                    "reply": reply,
                                    "chunks": chunks,
                                    "sources": sources,
                                    "types": types,
                                    "confidences": confidences,
                                    "top_conf": top_conf,
                                    "sender": sender,
                                    "subject": subject
                                }

                        # ─── Show reply + Send button ─────────────
                        if reply_key in st.session_state:
                            data = st.session_state[reply_key]

                            if data["top_conf"] < 40:
                                st.warning(f"⚠️ Low confidence ({data['top_conf']}%) — reply may be generic.")

                            st.subheader("📝 Draft Reply")

                            # Editable reply before sending
                            edited_reply = st.text_area(
                                "Review and edit before sending:",
                                value=data["reply"],
                                height=250,
                                key=f"edit_{email_id}"
                            )

                            col_send, col_discard = st.columns([1, 4])
                            with col_send:
                                if st.button(f"📤 Send Reply", key=f"send_{email_id}", type="primary"):
                                    with st.spinner("Sending reply..."):
                                        success = send_reply(
                                            access_token,
                                            email_id,
                                            edited_reply,
                                            data["sender"],
                                            data["subject"]
                                        )
                                        if success:
                                            mark_as_read(access_token, email_id)
                                            st.session_state[sent_key] = True
                                            del st.session_state[reply_key]
                                            st.success("✅ Reply sent and email marked as read!")
                                            st.rerun()
                                        else:
                                            st.error("❌ Failed to send. Check Graph API permissions.")
                            with col_discard:
                                if st.button(f"🗑️ Discard", key=f"discard_{email_id}"):
                                    del st.session_state[reply_key]
                                    st.rerun()

                            with st.container(border=True):
                                st.markdown("🔍 **Retrieved context**")
                                for chunk, src, t, conf in zip(data["chunks"], data["sources"], data["types"], data["confidences"]):
                                    st.markdown(f"**[{t.upper()}] {src}** — {confidence_badge(conf)}")
                                    st.markdown(chunk)
                                    st.divider()

        if st.button("🔓 Disconnect Outlook"):
            for key in ["access_token", "msal_flow", "msal_app", "oauth_started"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()

# ══════════════════════════════════════════════════════════════════
# TAB 2 — Manual Input
# ══════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("✍️ Paste Email Manually")
    new_email = st.text_area("Paste the incoming email here:", height=150,
        placeholder="e.g. Our MFT transfer to partner XYZ is failing with error code 1234...")

    if st.button("Generate Reply ✨", type="primary", key="manual_generate"):
        if new_email.strip():
            with st.spinner("Searching knowledge base..."):
                reply, chunks, sources, types, confidences, top_conf = generate_reply(new_email, collection)

            if top_conf < 40:
                st.warning(f"⚠️ Low confidence match ({top_conf}%) — reply may be generic.")

            st.subheader("📝 Draft Reply")
            st.markdown(reply)
            st.text_area("Copy from here:", value=reply, height=200, key="manual_copy")

            with st.expander("🔍 Retrieved context (what ChromaDB found)"):
                for chunk, src, t, conf in zip(chunks, sources, types, confidences):
                    st.markdown(f"**[{t.upper()}] Source: {src}** — {confidence_badge(conf)}")
                    st.markdown(chunk)
                    st.divider()
        else:
            st.warning("Please enter an email query first.")